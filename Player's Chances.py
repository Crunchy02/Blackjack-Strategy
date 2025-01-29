import numpy as np
import random
import time
import openpyxl

#Shuffles and cuts the deck and places a placeholder in it
def preparingDeck(deck):
    # Shuffling the deck
    random.shuffle(deck)

    # Cutting the deck
    #cutPoint = random.randint(50, 275)
    #deck = deck[cutPoint:] + deck[:cutPoint]

    # Inserting placeholder to indicate when to reshuffle deck
    placeholderLoc = random.randint(150, 250)
    deck = deck[:placeholderLoc] + [0] + deck[placeholderLoc:]

    return deck

if __name__ == '__main__':
    startTime = time.time()

    #Opening excel file to store stats in
    wb = openpyxl.load_workbook('Blackjack Data.xlsx')
    ws = wb['Player Stats']

    #Makes a deck consisting of 6 standard 52 card decks
    deck = [0 for i in range(52)]
    for normalCards in range(2,10):
        for numOfDuplicates in range(4):
             deck[numOfDuplicates + (normalCards - 2)*4] = normalCards
    for faceCards in range(1,17):
        deck[31 + faceCards] = 10
    for aces in range(1,5):
        deck[47 + aces] = 11

    deck = deck*6

    deck = preparingDeck(deck)

    #Initializing arrays to hold stats for both when the player hits 17-21 and busts
    playerHittingStats = [[0 for cols in range(20)] for rows in range(15)]
    playerBustingStats = [[0 for cols in range(20)] for rows in range(3)]

    columnTotals = [0 for j in range(20)]

    #The primary loop where the player draws 3 cards. Checks after each card drawn whether or not the player has hit 17-21
    for i in range(130000000):
        if (i % 1300000) == 0:
            print('|', end="")
        
        reshuffle = False

        #Checking to see if either of the next two cards are the placeholder card
        if deck[0] == 0:
            deck = deck[1:]
            reshuffle = True
        if deck[1] == 0:
            deck = [deck[0]] + deck[2:]
            reshuffle = True

        #Drawing the first two cards into the players hand (places them also at the end of the deck to be reused when deck is shuffled)
        playersHand = deck[:2]
        deck = deck[2:] + playersHand

        #Checking the player's starting hand score
        score = 0
        for j in range(2):
            score += playersHand[j]

        #If the player has 2 aces, both are changed to equal 1 in order to record stats of a starting hand of 2 (in case one were to split 2's)
        if score == 22:
            playersHand = [1, 1]
            score = 2

        #If the player has an ace and a 2, the ace is changed to equal 1 in order to record stats of a starting hand of 3 (in case one were to split 3's)
        if score == 13 and (playersHand[0] == 2 or playersHand[1] == 2):
            playersHand = [2, 1]
            score = 3

        #Recording total number of times each starting score occured
        column = score - 2
        columnTotals[column] += 1

        #Drawing the 3 cards and checking/recording score after each draw
        for cardsDrawn in range(3):
            if deck[0] == 0:
                deck = deck[1:]
                reshuffle = True

            playersHand.append(deck.pop(0))
            deck.append(playersHand[-1])

            #Checking if an ace needs to be changed to a 1
            if playersHand[-1] == 11 and score + playersHand[-1] > 21:
                playersHand[-1] = 1

            score += playersHand[-1]

            #Recording stats
            if score > 21:
                playerBustingStats[cardsDrawn][column] += 1
            elif score > 16:
                playerHittingStats[cardsDrawn * 5 + score - 17][column] += 1

            if reshuffle:
                deck = preparingDeck(deck)
                reshuffle = False

    #Turning stats into percentages and exporting to an excel file
    for cols in range(20):
        for rows in range(15):
            ws.cell(row=3 + rows + rows // 5 * 2, column=cols + 3).value = playerHittingStats[rows][cols] / columnTotals[cols] * 100

    for cols in range(20):
        for rows in range(3):
            ws.cell(row=3 + rows * 7, column=cols + 26).value = playerBustingStats[rows][cols] / columnTotals[cols] * 100

    wb.save('Blackjack Data.xlsx')

    print(columnTotals)
    print(time.time() - startTime)
